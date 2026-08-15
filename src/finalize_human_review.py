import json
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parent.parent

INPUT_FILE = (
    PROJECT_ROOT
    / "05_Evidence_Register"
    / "TJX_batch_evidence_review.xlsx"
)

OUTPUT_FOLDER = (
    PROJECT_ROOT
    / "05_Evidence_Register"
    / "reviewed_outputs"
)

APPROVED_OUTPUT_FILE = (
    OUTPUT_FOLDER
    / "TJX_approved_evidence.json"
)

REJECTED_OUTPUT_FILE = (
    OUTPUT_FOLDER
    / "TJX_rejected_evidence.json"
)

SUMMARY_OUTPUT_FILE = (
    OUTPUT_FOLDER
    / "TJX_human_review_summary.json"
)

SHEET_NAME = "Evidence Review"


def utc_timestamp() -> str:
    return datetime.now(timezone.utc).isoformat()


def load_review_rows() -> list[dict[str, Any]]:
    """
    Read all reviewed evidence rows from Excel.
    """

    if not INPUT_FILE.exists():
        raise FileNotFoundError(
            f"Review workbook not found:\n{INPUT_FILE}"
        )

    workbook = load_workbook(
        INPUT_FILE,
        data_only=True,
    )

    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(
            f"Worksheet '{SHEET_NAME}' was not found."
        )

    worksheet = workbook[SHEET_NAME]

    headers = [
        cell.value
        for cell in worksheet[1]
    ]

    required_headers = {
        "Evidence ID",
        "Source Document",
        "PDF Page",
        "Topic",
        "Claim",
        "Exact Quote",
        "Evidence Type",
        "Human Review Status",
        "Correction Needed",
        "Corrected Claim",
        "Reviewer Comment",
        "Duplicate Evidence",
    }

    missing_headers = required_headers - set(headers)

    if missing_headers:
        raise ValueError(
            f"Missing required columns: "
            f"{sorted(missing_headers)}"
        )

    rows: list[dict[str, Any]] = []

    for row_number in range(
        2,
        worksheet.max_row + 1,
    ):
        row_values = [
            worksheet.cell(
                row=row_number,
                column=column_number,
            ).value
            for column_number in range(
                1,
                worksheet.max_column + 1,
            )
        ]

        row_record = dict(
            zip(
                headers,
                row_values,
            )
        )

        evidence_id = row_record.get(
            "Evidence ID"
        )

        if not evidence_id:
            continue

        row_record["_excel_row_number"] = row_number

        rows.append(row_record)

    return rows


def validate_human_review(
    rows: list[dict[str, Any]],
) -> None:
    """
    Confirm that every row has a complete and internally consistent
    human-review decision.
    """

    allowed_statuses = {
        "Approved",
        "Approved with Correction",
        "Rejected",
    }

    allowed_yes_no = {
        "Yes",
        "No",
    }

    errors: list[str] = []

    for row in rows:
        evidence_id = row.get("Evidence ID")
        excel_row = row.get("_excel_row_number")
        status = row.get("Human Review Status")
        correction_needed = row.get(
            "Correction Needed"
        )
        corrected_claim = row.get(
            "Corrected Claim"
        )
        reviewer_comment = row.get(
            "Reviewer Comment"
        )
        duplicate_status = row.get(
            "Duplicate Evidence"
        )

        prefix = (
            f"Excel row {excel_row}, "
            f"{evidence_id}"
        )

        if status not in allowed_statuses:
            errors.append(
                f"{prefix}: invalid or incomplete "
                f"Human Review Status {status!r}."
            )

        if correction_needed not in allowed_yes_no:
            errors.append(
                f"{prefix}: Correction Needed must be "
                f"'Yes' or 'No'."
            )

        if duplicate_status not in allowed_yes_no:
            errors.append(
                f"{prefix}: Duplicate Evidence must be "
                f"'Yes' or 'No'."
            )

        if status == "Approved":
            if correction_needed == "Yes":
                errors.append(
                    f"{prefix}: an Approved item should not "
                    f"normally have Correction Needed = Yes."
                )

        if status == "Approved with Correction":
            if correction_needed != "Yes":
                errors.append(
                    f"{prefix}: Approved with Correction "
                    f"requires Correction Needed = Yes."
                )

            if not corrected_claim:
                errors.append(
                    f"{prefix}: Corrected Claim is required."
                )

        if status == "Rejected":
            if not reviewer_comment:
                errors.append(
                    f"{prefix}: a rejected item requires "
                    f"a Reviewer Comment."
                )

    if errors:
        print("\nHuman-review validation errors:")

        for error in errors:
            print(f"- {error}")

        raise ValueError(
            "Human review is incomplete or inconsistent."
        )


def build_final_claim(
    row: dict[str, Any],
) -> str:
    """
    Use the corrected claim when applicable; otherwise use the
    original AI claim.
    """

    status = row.get(
        "Human Review Status"
    )

    corrected_claim = row.get(
        "Corrected Claim"
    )

    if (
        status == "Approved with Correction"
        and corrected_claim
    ):
        return str(corrected_claim).strip()

    return str(
        row.get("Claim") or ""
    ).strip()


def clean_row_for_output(
    row: dict[str, Any],
) -> dict[str, Any]:
    """
    Remove temporary Excel-processing fields and add the final,
    human-approved claim.
    """

    cleaned = {
        key: value
        for key, value in row.items()
        if not key.startswith("_")
    }

    cleaned["Final Claim"] = build_final_claim(
        row
    )

    cleaned["Final Validation Status"] = (
        "human_validated"
        if row.get("Human Review Status")
        in {
            "Approved",
            "Approved with Correction",
        }
        else "human_rejected"
    )

    return cleaned


def split_reviewed_rows(
    rows: list[dict[str, Any]],
) -> tuple[
    list[dict[str, Any]],
    list[dict[str, Any]],
]:
    """
    Separate accepted evidence from rejected evidence.

    Duplicate items are retained but clearly flagged so they can be
    removed during later deduplication.
    """

    approved_rows: list[dict[str, Any]] = []
    rejected_rows: list[dict[str, Any]] = []

    for row in rows:
        cleaned_row = clean_row_for_output(
            row
        )

        status = row.get(
            "Human Review Status"
        )

        if status in {
            "Approved",
            "Approved with Correction",
        }:
            approved_rows.append(
                cleaned_row
            )

        elif status == "Rejected":
            rejected_rows.append(
                cleaned_row
            )

    return approved_rows, rejected_rows


def calculate_summary(
    rows: list[dict[str, Any]],
    approved_rows: list[dict[str, Any]],
    rejected_rows: list[dict[str, Any]],
) -> dict[str, Any]:
    """
    Calculate quality and review metrics.
    """

    total = len(rows)

    status_counts = Counter(
        row.get("Human Review Status")
        for row in rows
    )

    evidence_type_counts = Counter(
        row.get("Evidence Type")
        for row in approved_rows
        if row.get("Evidence Type")
    )

    topic_counts = Counter(
        row.get("Topic")
        for row in approved_rows
        if row.get("Topic")
    )

    page_counts = Counter(
        row.get("PDF Page")
        for row in approved_rows
    )

    duplicates = sum(
        row.get("Duplicate Evidence") == "Yes"
        for row in approved_rows
    )

    approved_without_correction = (
        status_counts.get(
            "Approved",
            0,
        )
    )

    approved_with_correction = (
        status_counts.get(
            "Approved with Correction",
            0,
        )
    )

    rejected = status_counts.get(
        "Rejected",
        0,
    )

    accepted_total = len(
        approved_rows
    )

    return {
        "generated_at_utc": utc_timestamp(),
        "source_workbook": str(INPUT_FILE),
        "total_evidence_items": total,
        "approved_without_correction": (
            approved_without_correction
        ),
        "approved_with_correction": (
            approved_with_correction
        ),
        "rejected": rejected,
        "accepted_total": accepted_total,
        "overall_acceptance_rate": (
            accepted_total / total
            if total
            else 0
        ),
        "clean_approval_rate": (
            approved_without_correction / total
            if total
            else 0
        ),
        "correction_rate": (
            approved_with_correction / total
            if total
            else 0
        ),
        "rejection_rate": (
            rejected / total
            if total
            else 0
        ),
        "approved_duplicate_items": duplicates,
        "approved_nonduplicate_items": (
            accepted_total - duplicates
        ),
        "approved_evidence_by_type": dict(
            evidence_type_counts
        ),
        "approved_evidence_by_topic": dict(
            topic_counts
        ),
        "approved_evidence_by_page": {
            str(page): count
            for page, count
            in sorted(page_counts.items())
        },
        "review_complete": True,
    }


def save_json(
    output_path: Path,
    data: Any,
) -> None:
    """
    Save JSON with readable UTF-8 formatting.
    """

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    with output_path.open(
        "w",
        encoding="utf-8",
    ) as file:
        json.dump(
            data,
            file,
            ensure_ascii=False,
            indent=2,
        )


def print_summary(
    summary: dict[str, Any],
) -> None:
    print("=" * 72)
    print("TJX HUMAN-REVIEW FINALIZATION SUMMARY")
    print("=" * 72)

    print(
        f"Total evidence items: "
        f"{summary['total_evidence_items']}"
    )

    print(
        f"Approved without correction: "
        f"{summary['approved_without_correction']}"
    )

    print(
        f"Approved with correction: "
        f"{summary['approved_with_correction']}"
    )

    print(
        f"Rejected: "
        f"{summary['rejected']}"
    )

    print(
        f"Approved duplicate items: "
        f"{summary['approved_duplicate_items']}"
    )

    print("-" * 72)

    print(
        f"Overall acceptance rate: "
        f"{summary['overall_acceptance_rate']:.1%}"
    )

    print(
        f"Clean approval rate: "
        f"{summary['clean_approval_rate']:.1%}"
    )

    print(
        f"Correction rate: "
        f"{summary['correction_rate']:.1%}"
    )

    print(
        f"Rejection rate: "
        f"{summary['rejection_rate']:.1%}"
    )

    print("-" * 72)

    print(
        f"Approved evidence saved to:\n"
        f"{APPROVED_OUTPUT_FILE}"
    )

    print(
        f"\nRejected evidence saved to:\n"
        f"{REJECTED_OUTPUT_FILE}"
    )

    print(
        f"\nReview summary saved to:\n"
        f"{SUMMARY_OUTPUT_FILE}"
    )

    print("=" * 72)


def main() -> None:
    rows = load_review_rows()

    if not rows:
        raise ValueError(
            "No evidence rows were found in the workbook."
        )

    validate_human_review(
        rows
    )

    approved_rows, rejected_rows = (
        split_reviewed_rows(
            rows
        )
    )

    summary = calculate_summary(
        rows=rows,
        approved_rows=approved_rows,
        rejected_rows=rejected_rows,
    )

    save_json(
        APPROVED_OUTPUT_FILE,
        approved_rows,
    )

    save_json(
        REJECTED_OUTPUT_FILE,
        rejected_rows,
    )

    save_json(
        SUMMARY_OUTPUT_FILE,
        summary,
    )

    print_summary(
        summary
    )


if __name__ == "__main__":
    main()