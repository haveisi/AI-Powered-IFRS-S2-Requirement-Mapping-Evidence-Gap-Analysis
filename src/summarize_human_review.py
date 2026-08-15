from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parent.parent

INPUT_FILE = (
    PROJECT_ROOT
    / "05_Evidence_Register"
    / "test_evidence_review.xlsx"
)

SHEET_NAME = "Evidence Review"


def load_review_rows() -> list[dict[str, Any]]:
    """
    Load human-review decisions from the Excel workbook.
    """

    if not INPUT_FILE.exists():
        raise FileNotFoundError(
            f"Review workbook not found:\n{INPUT_FILE}"
        )

    workbook = load_workbook(
        INPUT_FILE,
        data_only=True,
    )

    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(
            f"Worksheet '{SHEET_NAME}' was not found."
        )

    worksheet = workbook[SHEET_NAME]

    headers = [
        cell.value
        for cell in worksheet[1]
    ]

    required_headers = {
        "Evidence ID",
        "Claim",
        "Exact Quote",
        "Human Review Status",
        "Reviewer Comment",
        "Correction Needed",
        "Corrected Claim",
    }

    missing_headers = required_headers - set(headers)

    if missing_headers:
        raise ValueError(
            f"Missing required columns: "
            f"{sorted(missing_headers)}"
        )

    rows: list[dict[str, Any]] = []

    for row_number in range(
        2,
        worksheet.max_row + 1,
    ):
        values = [
            worksheet.cell(
                row=row_number,
                column=column_number,
            ).value
            for column_number in range(
                1,
                worksheet.max_column + 1,
            )
        ]

        row_record = dict(
            zip(
                headers,
                values,
            )
        )

        evidence_id = row_record.get(
            "Evidence ID"
        )

        if not evidence_id:
            continue

        rows.append(row_record)

    return rows


def validate_review_decisions(
    rows: list[dict[str, Any]],
) -> None:
    """
    Confirm that every evidence item has a valid review status.
    """

    allowed_statuses = {
        "Approved",
        "Approved with Correction",
        "Rejected",
    }

    errors: list[str] = []

    for row in rows:
        evidence_id = row.get("Evidence ID")
        status = row.get("Human Review Status")
        correction_needed = row.get(
            "Correction Needed"
        )
        corrected_claim = row.get(
            "Corrected Claim"
        )

        if status not in allowed_statuses:
            errors.append(
                f"{evidence_id}: invalid or incomplete status "
                f"{status!r}"
            )

        if (
            status == "Approved with Correction"
            and correction_needed != "Yes"
        ):
            errors.append(
                f"{evidence_id}: status is "
                f"'Approved with Correction' but "
                f"'Correction Needed' is not 'Yes'."
            )

        if (
            status == "Approved with Correction"
            and not corrected_claim
        ):
            errors.append(
                f"{evidence_id}: corrected claim is required."
            )

        if (
            status == "Rejected"
            and not row.get("Reviewer Comment")
        ):
            errors.append(
                f"{evidence_id}: rejected items should include "
                f"a reviewer comment."
            )

    if errors:
        print("\nReview validation errors:")

        for error in errors:
            print(f"- {error}")

        raise ValueError(
            "Human review is incomplete or inconsistent."
        )


def calculate_review_metrics(
    rows: list[dict[str, Any]],
) -> dict[str, Any]:
    """
    Calculate human-review results.
    """

    total = len(rows)

    approved = sum(
        row.get("Human Review Status") == "Approved"
        for row in rows
    )

    approved_with_correction = sum(
        row.get("Human Review Status")
        == "Approved with Correction"
        for row in rows
    )

    rejected = sum(
        row.get("Human Review Status") == "Rejected"
        for row in rows
    )

    accepted_total = (
        approved
        + approved_with_correction
    )

    approval_rate = (
        accepted_total / total
        if total
        else 0
    )

    clean_approval_rate = (
        approved / total
        if total
        else 0
    )

    correction_rate = (
        approved_with_correction / total
        if total
        else 0
    )

    rejection_rate = (
        rejected / total
        if total
        else 0
    )

    return {
        "total": total,
        "approved": approved,
        "approved_with_correction": (
            approved_with_correction
        ),
        "rejected": rejected,
        "accepted_total": accepted_total,
        "approval_rate": approval_rate,
        "clean_approval_rate": clean_approval_rate,
        "correction_rate": correction_rate,
        "rejection_rate": rejection_rate,
    }


def print_review_summary(
    metrics: dict[str, Any],
) -> None:
    """
    Print a readable review summary.
    """

    print("=" * 65)
    print("TJX TEST EVIDENCE — HUMAN REVIEW SUMMARY")
    print("=" * 65)

    print(
        f"Total evidence items: "
        f"{metrics['total']}"
    )

    print(
        f"Approved without correction: "
        f"{metrics['approved']}"
    )

    print(
        f"Approved with correction: "
        f"{metrics['approved_with_correction']}"
    )

    print(
        f"Rejected: "
        f"{metrics['rejected']}"
    )

    print("-" * 65)

    print(
        f"Overall acceptance rate: "
        f"{metrics['approval_rate']:.1%}"
    )

    print(
        f"Clean approval rate: "
        f"{metrics['clean_approval_rate']:.1%}"
    )

    print(
        f"Correction rate: "
        f"{metrics['correction_rate']:.1%}"
    )

    print(
        f"Rejection rate: "
        f"{metrics['rejection_rate']:.1%}"
    )

    print("=" * 65)


def print_item_results(
    rows: list[dict[str, Any]],
) -> None:
    """
    Print the decision for each evidence item.
    """

    print("\nItem-level decisions:")

    for row in rows:
        evidence_id = row.get(
            "Evidence ID"
        )

        status = row.get(
            "Human Review Status"
        )

        topic = row.get(
            "Topic",
            "",
        )

        print(
            f"- {evidence_id}: "
            f"{status} | {topic}"
        )

        if status == "Approved with Correction":
            print(
                f"  Corrected claim: "
                f"{row.get('Corrected Claim')}"
            )

        reviewer_comment = row.get(
            "Reviewer Comment"
        )

        if reviewer_comment:
            print(
                f"  Reviewer comment: "
                f"{reviewer_comment}"
            )


def main() -> None:
    rows = load_review_rows()

    if not rows:
        raise ValueError(
            "No evidence rows were found."
        )

    validate_review_decisions(rows)

    metrics = calculate_review_metrics(rows)

    print_review_summary(metrics)

    print_item_results(rows)


if __name__ == "__main__":
    main()