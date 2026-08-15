import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


PROJECT_ROOT = Path(__file__).resolve().parent.parent

INPUT_FOLDER = (
    PROJECT_ROOT
    / "05_Evidence_Register"
    / "batch_pages"
)

OUTPUT_FILE = (
    PROJECT_ROOT
    / "05_Evidence_Register"
    / "TJX_batch_evidence_review.xlsx"
)


def load_page_results() -> list[dict[str, Any]]:
    """
    Load all validated page-level evidence JSON files.
    """

    if not INPUT_FOLDER.exists():
        raise FileNotFoundError(
            f"Input folder not found:\n{INPUT_FOLDER}"
        )

    json_files = sorted(
        INPUT_FOLDER.glob("page_*_evidence.json")
    )

    if not json_files:
        raise FileNotFoundError(
            f"No page evidence files found in:\n{INPUT_FOLDER}"
        )

    results: list[dict[str, Any]] = []

    for json_path in json_files:
        with json_path.open(
            "r",
            encoding="utf-8",
        ) as file:
            result = json.load(file)

        if not isinstance(result, dict):
            raise ValueError(
                f"{json_path.name} does not contain a JSON object."
            )

        evidence_items = result.get("evidence_items")

        if not isinstance(evidence_items, list):
            raise ValueError(
                f"{json_path.name} has no valid evidence_items list."
            )

        result["_source_json_file"] = json_path.name
        results.append(result)

    return results


def flatten_evidence(
    page_results: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """
    Convert page-level JSON records into one row per evidence item.
    """

    rows: list[dict[str, Any]] = []

    for page_result in page_results:
        source_document = page_result.get(
            "source_document"
        )

        pdf_page_number = page_result.get(
            "pdf_page_number"
        )

        model_name = page_result.get(
            "model_name"
        )

        extracted_at = page_result.get(
            "extracted_at_utc"
        )

        selection_topic = page_result.get(
            "selection_primary_topic",
            page_result.get("page_primary_topic"),
        )

        selection_score = page_result.get(
            "selection_relevance_score"
        )

        api_usage = page_result.get(
            "api_usage",
            {},
        )

        for item in page_result.get(
            "evidence_items",
            [],
        ):
            rows.append(
                {
                    "Evidence ID": item.get("evidence_id"),
                    "Source Document": source_document,
                    "PDF Page": pdf_page_number,
                    "Selection Topic": selection_topic,
                    "Selection Score": selection_score,
                    "Topic": item.get("topic"),
                    "Claim": item.get("claim"),
                    "Exact Quote": item.get("exact_quote"),
                    "Evidence Type": item.get("evidence_type"),
                    "Metric Name": item.get("metric_name"),
                    "Metric Value": item.get("metric_value"),
                    "Metric Unit": item.get("metric_unit"),
                    "Reporting Period": item.get("reporting_period"),
                    "Geographic Scope": item.get("geographic_scope"),
                    "AI Confidence": item.get("confidence"),
                    "Quote Validation": item.get(
                        "quote_validation_status"
                    ),
                    "Human Review Status": "Pending",
                    "Correction Needed": "No",
                    "Corrected Claim": "",
                    "Reviewer Comment": "",
                    "Framework Relevance": "",
                    "Potential Disclosure Requirement": "",
                    "Duplicate Evidence": "No",
                    "Source JSON File": page_result.get(
                        "_source_json_file"
                    ),
                    "Model Name": model_name,
                    "Input Tokens": api_usage.get(
                        "input_tokens"
                    ),
                    "Output Tokens": api_usage.get(
                        "output_tokens"
                    ),
                    "Extracted At UTC": extracted_at,
                }
            )

    return rows


def create_workbook(
    rows: list[dict[str, Any]],
) -> Workbook:
    """
    Create the consolidated human-review workbook.
    """

    workbook = Workbook()

    review_sheet = workbook.active
    review_sheet.title = "Evidence Review"

    headers = list(rows[0].keys())

    review_sheet.append(headers)

    for row in rows:
        review_sheet.append(
            [row.get(header) for header in headers]
        )

    format_review_sheet(
        review_sheet,
        headers,
    )

    add_review_dropdowns(
        review_sheet,
        headers,
    )

    create_summary_sheet(
        workbook,
        rows,
    )

    create_instructions_sheet(
        workbook
    )

    return workbook


def format_review_sheet(
    worksheet,
    headers: list[str],
) -> None:
    """
    Apply readable formatting.
    """

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    widths = {
        "Evidence ID": 16,
        "Source Document": 42,
        "PDF Page": 10,
        "Selection Topic": 24,
        "Selection Score": 14,
        "Topic": 24,
        "Claim": 55,
        "Exact Quote": 75,
        "Evidence Type": 18,
        "Metric Name": 22,
        "Metric Value": 16,
        "Metric Unit": 16,
        "Reporting Period": 18,
        "Geographic Scope": 22,
        "AI Confidence": 14,
        "Quote Validation": 18,
        "Human Review Status": 24,
        "Correction Needed": 18,
        "Corrected Claim": 55,
        "Reviewer Comment": 45,
        "Framework Relevance": 26,
        "Potential Disclosure Requirement": 38,
        "Duplicate Evidence": 18,
        "Source JSON File": 28,
        "Model Name": 24,
        "Input Tokens": 14,
        "Output Tokens": 14,
        "Extracted At UTC": 24,
    }

    for column_number, header in enumerate(
        headers,
        start=1,
    ):
        column_letter = worksheet.cell(
            row=1,
            column=column_number,
        ).column_letter

        worksheet.column_dimensions[
            column_letter
        ].width = widths.get(
            header,
            18,
        )

    for row in worksheet.iter_rows(
        min_row=2,
        max_row=worksheet.max_row,
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    for row_number in range(
        2,
        worksheet.max_row + 1,
    ):
        worksheet.row_dimensions[
            row_number
        ].height = 95


def get_column_letter_by_header(
    headers: list[str],
    header_name: str,
) -> str:
    """
    Return Excel column letter for a header.
    """

    if header_name not in headers:
        raise ValueError(
            f"Header not found: {header_name}"
        )

    column_number = (
        headers.index(header_name) + 1
    )

    return chr(
        64 + column_number
    ) if column_number <= 26 else None


def add_review_dropdowns(
    worksheet,
    headers: list[str],
) -> None:
    """
    Add controlled review dropdowns.
    """

    review_status = DataValidation(
        type="list",
        formula1=(
            '"Pending,Approved,'
            'Approved with Correction,Rejected"'
        ),
        allow_blank=False,
    )

    yes_no = DataValidation(
        type="list",
        formula1='"No,Yes"',
        allow_blank=False,
    )

    framework = DataValidation(
        type="list",
        formula1=(
            '"Not Assessed,ESRS,ISSB,California,'
            'EU Taxonomy,Multiple,None"'
        ),
        allow_blank=True,
    )

    worksheet.add_data_validation(
        review_status
    )

    worksheet.add_data_validation(
        yes_no
    )

    worksheet.add_data_validation(
        framework
    )

    review_column = get_column_letter_by_header(
        headers,
        "Human Review Status",
    )

    correction_column = get_column_letter_by_header(
        headers,
        "Correction Needed",
    )

    duplicate_column = get_column_letter_by_header(
        headers,
        "Duplicate Evidence",
    )

    framework_column = get_column_letter_by_header(
        headers,
        "Framework Relevance",
    )

    last_row = worksheet.max_row

    review_status.add(
        f"{review_column}2:{review_column}{last_row}"
    )

    yes_no.add(
        f"{correction_column}2:{correction_column}{last_row}"
    )

    yes_no.add(
        f"{duplicate_column}2:{duplicate_column}{last_row}"
    )

    framework.add(
        f"{framework_column}2:{framework_column}{last_row}"
    )


def create_summary_sheet(
    workbook: Workbook,
    rows: list[dict[str, Any]],
) -> None:
    """
    Create a basic batch summary.
    """

    worksheet = workbook.create_sheet(
        "Batch Summary"
    )

    page_numbers = sorted(
        {
            row["PDF Page"]
            for row in rows
        }
    )

    total_input_tokens = sum(
        row.get("Input Tokens") or 0
        for row in rows
    )

    total_output_tokens = sum(
        row.get("Output Tokens") or 0
        for row in rows
    )

    summary_rows = [
        ["Metric", "Value"],
        ["Evidence items", len(rows)],
        ["Pages represented", len(page_numbers)],
        [
            "PDF pages",
            ", ".join(
                str(page)
                for page in page_numbers
            ),
        ],
        [
            "Distinct topics",
            len(
                {
                    row["Topic"]
                    for row in rows
                    if row.get("Topic")
                }
            ),
        ],
        [
            "Total input tokens",
            total_input_tokens,
        ],
        [
            "Total output tokens",
            total_output_tokens,
        ],
        [
            "Human review status",
            "Pending",
        ],
    ]

    for row in summary_rows:
        worksheet.append(row)

    worksheet.column_dimensions["A"].width = 30
    worksheet.column_dimensions["B"].width = 60

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )


def create_instructions_sheet(
    workbook: Workbook,
) -> None:
    """
    Create human-review instructions.
    """

    worksheet = workbook.create_sheet(
        "Review Instructions"
    )

    instructions = [
        ["Field", "Review instruction"],
        [
            "Claim",
            "Confirm that the claim does not exaggerate or omit important context from the quote.",
        ],
        [
            "Exact Quote",
            "Confirm that the quotation directly supports the claim.",
        ],
        [
            "Evidence Type",
            "Confirm whether the item is a policy, target, metric, action, governance statement, risk, or other evidence.",
        ],
        [
            "Human Review Status",
            "Choose Approved, Approved with Correction, or Rejected.",
        ],
        [
            "Correction Needed",
            "Set to Yes when the claim or classification requires revision.",
        ],
        [
            "Corrected Claim",
            "Enter a revised claim when approving with correction.",
        ],
        [
            "Reviewer Comment",
            "Explain rejected items, missing context, or classification concerns.",
        ],
        [
            "Framework Relevance",
            "Do not guess. Select a framework only when the evidence may support a specific disclosure requirement.",
        ],
        [
            "Potential Disclosure Requirement",
            "Leave blank until the framework library is built and reviewed.",
        ],
        [
            "Duplicate Evidence",
            "Set to Yes when substantially the same evidence appears elsewhere.",
        ],
    ]

    for row in instructions:
        worksheet.append(row)

    worksheet.column_dimensions["A"].width = 34
    worksheet.column_dimensions["B"].width = 100

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )


def save_workbook(
    workbook: Workbook,
) -> None:
    OUTPUT_FILE.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook.save(
        OUTPUT_FILE
    )


def main() -> None:
    page_results = load_page_results()

    rows = flatten_evidence(
        page_results
    )

    if not rows:
        raise ValueError(
            "No evidence items were found."
        )

    workbook = create_workbook(
        rows
    )

    save_workbook(
        workbook
    )

    print("=" * 70)
    print("TJX CONSOLIDATED BATCH EVIDENCE REGISTER")
    print("=" * 70)
    print(
        f"Page files loaded: {len(page_results)}"
    )
    print(
        f"Evidence items consolidated: {len(rows)}"
    )
    print(
        f"Workbook saved to:\n{OUTPUT_FILE}"
    )
    print("=" * 70)


if __name__ == "__main__":
    main()