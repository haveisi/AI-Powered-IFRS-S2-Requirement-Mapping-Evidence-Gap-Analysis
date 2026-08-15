import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


PROJECT_ROOT = Path(__file__).resolve().parent.parent

INPUT_FILE = (
    PROJECT_ROOT
    / "05_Evidence_Register"
    / "test_evidence.json"
)

OUTPUT_FILE = (
    PROJECT_ROOT
    / "05_Evidence_Register"
    / "test_evidence_review.xlsx"
)


def load_evidence() -> dict[str, Any]:
    """
    Load the validated evidence JSON.
    """

    if not INPUT_FILE.exists():
        raise FileNotFoundError(
            f"Evidence JSON not found:\n{INPUT_FILE}"
        )

    with INPUT_FILE.open("r", encoding="utf-8") as file:
        result = json.load(file)

    if not isinstance(result, dict):
        raise ValueError(
            "The evidence JSON must contain one JSON object."
        )

    evidence_items = result.get("evidence_items")

    if not isinstance(evidence_items, list):
        raise ValueError(
            "The evidence JSON must contain an evidence_items list."
        )

    return result


def create_review_workbook(
    result: dict[str, Any],
) -> Workbook:
    """
    Create an Excel workbook for human review.
    """

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Evidence Review"

    headers = [
        "Evidence ID",
        "Source Document",
        "PDF Page",
        "Topic",
        "Claim",
        "Exact Quote",
        "Evidence Type",
        "Metric Name",
        "Metric Value",
        "Metric Unit",
        "Reporting Period",
        "Geographic Scope",
        "AI Confidence",
        "Quote Validation",
        "Human Review Status",
        "Reviewer Comment",
        "Correction Needed",
        "Corrected Claim",
    ]

    worksheet.append(headers)

    source_document = result.get("source_document")
    pdf_page_number = result.get("pdf_page_number")

    for item in result.get("evidence_items", []):
        worksheet.append(
            [
                item.get("evidence_id"),
                source_document,
                pdf_page_number,
                item.get("topic"),
                item.get("claim"),
                item.get("exact_quote"),
                item.get("evidence_type"),
                item.get("metric_name"),
                item.get("metric_value"),
                item.get("metric_unit"),
                item.get("reporting_period"),
                item.get("geographic_scope"),
                item.get("confidence"),
                item.get("quote_validation_status"),
                "Pending",
                "",
                "No",
                "",
            ]
        )

    apply_formatting(worksheet)
    add_dropdown_controls(worksheet)

    create_instructions_sheet(workbook)

    return workbook


def apply_formatting(worksheet) -> None:
    """
    Apply readable formatting to the evidence-review sheet.
    """

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    column_widths = {
        "A": 14,
        "B": 42,
        "C": 12,
        "D": 24,
        "E": 55,
        "F": 75,
        "G": 18,
        "H": 22,
        "I": 16,
        "J": 16,
        "K": 18,
        "L": 22,
        "M": 15,
        "N": 18,
        "O": 20,
        "P": 40,
        "Q": 18,
        "R": 55,
    }

    for column_letter, width in column_widths.items():
        worksheet.column_dimensions[column_letter].width = width

    for row in worksheet.iter_rows(
        min_row=2,
        max_row=worksheet.max_row,
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    for row_number in range(
        2,
        worksheet.max_row + 1,
    ):
        worksheet.row_dimensions[row_number].height = 90


def add_dropdown_controls(worksheet) -> None:
    """
    Add controlled review-status dropdowns.
    """

    review_validation = DataValidation(
        type="list",
        formula1='"Pending,Approved,Approved with Correction,Rejected"',
        allow_blank=False,
    )

    correction_validation = DataValidation(
        type="list",
        formula1='"No,Yes"',
        allow_blank=False,
    )

    worksheet.add_data_validation(review_validation)
    worksheet.add_data_validation(correction_validation)

    review_validation.add(
        f"O2:O{worksheet.max_row}"
    )

    correction_validation.add(
        f"Q2:Q{worksheet.max_row}"
    )


def create_instructions_sheet(
    workbook: Workbook,
) -> None:
    """
    Create reviewer instructions.
    """

    worksheet = workbook.create_sheet(
        "Review Instructions"
    )

    instructions = [
        ["Field", "Review question"],
        [
            "Claim",
            "Does the claim faithfully summarize the quoted source text?",
        ],
        [
            "Exact Quote",
            "Does the quote directly support the claim without missing context?",
        ],
        [
            "Topic",
            "Is the ESG topic specific and accurate?",
        ],
        [
            "Evidence Type",
            "Is the item correctly classified as policy, target, metric, action, governance, risk, or other?",
        ],
        [
            "Metric Fields",
            "Are metric name, value, unit, period, and geography supported explicitly by the quote?",
        ],
        [
            "Approved",
            "Use only when the item is accurate and complete.",
        ],
        [
            "Approved with Correction",
            "Use when the evidence is valid but the claim or classification needs revision.",
        ],
        [
            "Rejected",
            "Use when the quote does not support the claim or the evidence is not useful.",
        ],
    ]

    for row in instructions:
        worksheet.append(row)

    worksheet.column_dimensions["A"].width = 28
    worksheet.column_dimensions["B"].width = 95

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )


def save_workbook(
    workbook: Workbook,
) -> None:
    """
    Save the Excel review workbook.
    """

    OUTPUT_FILE.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook.save(OUTPUT_FILE)


def main() -> None:
    result = load_evidence()

    workbook = create_review_workbook(
        result
    )

    save_workbook(
        workbook
    )

    evidence_count = len(
        result.get("evidence_items", [])
    )

    print(
        f"Evidence items exported: {evidence_count}"
    )

    print(
        f"Review workbook saved to:\n{OUTPUT_FILE}"
    )


if __name__ == "__main__":
    main()
