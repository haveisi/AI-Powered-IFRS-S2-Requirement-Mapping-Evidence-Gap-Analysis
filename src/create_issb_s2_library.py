from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


PROJECT_ROOT = Path(__file__).resolve().parent.parent

OUTPUT_FILE = (
    PROJECT_ROOT
    / "06_Framework_Libraries"
    / "ISSB_S2_Requirement_Library.xlsx"
)


HEADERS = [
    "Requirement ID",
    "Framework",
    "Standard",
    "Content Area",
    "Requirement Topic",
    "Requirement Summary",
    "Required Data Type",
    "Quantitative Required",
    "Time Horizon Required",
    "Financial Link Required",
    "Scenario Analysis Relevant",
    "Scope 1 Relevant",
    "Scope 2 Relevant",
    "Scope 3 Relevant",
    "Evidence Needed",
    "Source Reference",
    "Source URL",
    "Version",
    "Issued Date",
    "Effective Date",
    "Early Application Allowed",
    "Superseded",
    "Applicability Notes",
    "Review Status",
    "Reviewer Notes",
]


SOURCE_URL = (
    "https://www.ifrs.org/issued-standards/"
    "ifrs-sustainability-standards-navigator/"
    "ifrs-s2-climate-related-disclosures/"
)


def make_requirement(
    requirement_id: str,
    content_area: str,
    topic: str,
    summary: str,
    data_type: str,
    evidence_needed: str,
    source_reference: str,
    quantitative: str = "No",
    time_horizon: str = "No",
    financial_link: str = "No",
    scenario_analysis: str = "No",
    scope_1: str = "No",
    scope_2: str = "No",
    scope_3: str = "No",
    applicability_notes: str = "",
) -> dict:
    """
    Create one standardized IFRS S2 requirement record.
    """

    return {
        "Requirement ID": requirement_id,
        "Framework": "ISSB",
        "Standard": "IFRS S2",
        "Content Area": content_area,
        "Requirement Topic": topic,
        "Requirement Summary": summary,
        "Required Data Type": data_type,
        "Quantitative Required": quantitative,
        "Time Horizon Required": time_horizon,
        "Financial Link Required": financial_link,
        "Scenario Analysis Relevant": scenario_analysis,
        "Scope 1 Relevant": scope_1,
        "Scope 2 Relevant": scope_2,
        "Scope 3 Relevant": scope_3,
        "Evidence Needed": evidence_needed,
        "Source Reference": source_reference,
        "Source URL": SOURCE_URL,
        "Version": "IFRS S2 issued June 2023",
        "Issued Date": "2023-06-26",
        "Effective Date": "2024-01-01",
        "Early Application Allowed": "Yes",
        "Superseded": "No",
        "Applicability Notes": applicability_notes,
        "Review Status": "Draft",
        "Reviewer Notes": "",
    }


def build_requirements() -> list[dict]:
    """
    Create a controlled starter library.

    These are plain-language readiness summaries and should later
    be verified against the official IFRS S2 text.
    """

    return [
        make_requirement(
            requirement_id="S2-GOV-01",
            content_area="Governance",
            topic="Governance processes and controls",
            summary=(
                "Describe the governance processes, controls, and procedures "
                "used to oversee climate-related risks and opportunities."
            ),
            data_type="Governance",
            evidence_needed=(
                "Governance policy, committee charter, oversight process, "
                "responsibility statement, or documented climate control."
            ),
            source_reference="IFRS S2 paragraphs 5–7",
        ),

        make_requirement(
            requirement_id="S2-GOV-02",
            content_area="Governance",
            topic="Board oversight",
            summary=(
                "Identify the board, committee, or equivalent governing body "
                "responsible for climate-related oversight."
            ),
            data_type="Governance",
            evidence_needed=(
                "Board or committee name, charter, responsibility statement, "
                "or oversight disclosure."
            ),
            source_reference="IFRS S2 paragraph 6(a)",
        ),

        make_requirement(
            requirement_id="S2-GOV-03",
            content_area="Governance",
            topic="Management responsibilities",
            summary=(
                "Describe management's role in monitoring and managing "
                "climate-related risks and opportunities."
            ),
            data_type="Governance",
            evidence_needed=(
                "Named management roles, committee membership, reporting "
                "lines, delegated responsibilities, and control ownership."
            ),
            source_reference="IFRS S2 paragraph 6(b)",
        ),

        make_requirement(
            requirement_id="S2-GOV-04",
            content_area="Governance",
            topic="Oversight of climate targets",
            summary=(
                "Describe how the governing body oversees climate-related "
                "targets and monitors progress."
            ),
            data_type="Governance and target",
            evidence_needed=(
                "Board reporting, target dashboard, performance review, "
                "assigned accountability, or escalation procedures."
            ),
            source_reference="IFRS S2 paragraph 6(a)",
            quantitative="Yes",
        ),

        make_requirement(
            requirement_id="S2-STR-01",
            content_area="Strategy",
            topic="Climate risks and opportunities",
            summary=(
                "Identify climate-related risks and opportunities that could "
                "reasonably affect the entity's prospects."
            ),
            data_type="Risk and opportunity",
            evidence_needed=(
                "Climate risk register, opportunity register, materiality "
                "assessment, or documented identification process."
            ),
            source_reference="IFRS S2 paragraphs 8–12",
            time_horizon="Yes",
        ),

        make_requirement(
            requirement_id="S2-STR-02",
            content_area="Strategy",
            topic="Physical and transition risks",
            summary=(
                "Classify significant climate-related risks as physical "
                "risks or transition risks."
            ),
            data_type="Risk",
            evidence_needed=(
                "Risk classification, physical-risk analysis, transition-risk "
                "analysis, or climate risk register."
            ),
            source_reference="IFRS S2 paragraphs 10–12",
            time_horizon="Yes",
        ),

        make_requirement(
            requirement_id="S2-STR-03",
            content_area="Strategy",
            topic="Business model and value chain",
            summary=(
                "Describe current and anticipated effects of climate-related "
                "risks and opportunities on the business model and value chain."
            ),
            data_type="Strategy and value chain",
            evidence_needed=(
                "Supplier, sourcing, logistics, facility, store, product, "
                "customer, or business-model exposure analysis."
            ),
            source_reference="IFRS S2 paragraph 13",
            time_horizon="Yes",
            financial_link="Yes",
        ),

        make_requirement(
            requirement_id="S2-STR-04",
            content_area="Strategy",
            topic="Transition plan",
            summary=(
                "Describe relevant aspects of the climate transition plan, "
                "including assumptions, dependencies, actions, and milestones."
            ),
            data_type="Strategy, target, and action",
            evidence_needed=(
                "Decarbonization pathway, transition plan, milestones, "
                "implementation owners, assumptions, and dependencies."
            ),
            source_reference="IFRS S2 paragraph 14",
            quantitative="Yes",
            time_horizon="Yes",
            financial_link="Yes",
        ),

        make_requirement(
            requirement_id="S2-STR-05",
            content_area="Strategy",
            topic="Current financial effects",
            summary=(
                "Describe current financial effects of climate-related risks "
                "and opportunities on financial position, performance, and cash flows."
            ),
            data_type="Financial effect",
            evidence_needed=(
                "Climate-related revenue, cost, asset, liability, impairment, "
                "provision, capital expenditure, or cash-flow evidence."
            ),
            source_reference="IFRS S2 paragraphs 15–21",
            quantitative="Yes",
            financial_link="Yes",
        ),

        make_requirement(
            requirement_id="S2-STR-06",
            content_area="Strategy",
            topic="Climate resilience and scenario analysis",
            summary=(
                "Explain the resilience of the strategy and business model "
                "using climate-related scenario analysis."
            ),
            data_type="Scenario analysis and strategy",
            evidence_needed=(
                "Climate scenarios, assumptions, time horizons, vulnerabilities, "
                "resilience conclusions, and planned responses."
            ),
            source_reference="IFRS S2 paragraph 22",
            quantitative="Yes",
            time_horizon="Yes",
            financial_link="Yes",
            scenario_analysis="Yes",
        ),

        make_requirement(
            requirement_id="S2-RM-01",
            content_area="Risk Management",
            topic="Risk identification and assessment",
            summary=(
                "Describe processes used to identify and assess "
                "climate-related risks."
            ),
            data_type="Risk-management process",
            evidence_needed=(
                "Risk procedures, screening criteria, likelihood and severity "
                "methods, risk register, and assessment documentation."
            ),
            source_reference="IFRS S2 paragraphs 24–26",
            time_horizon="Yes",
        ),

        make_requirement(
            requirement_id="S2-RM-02",
            content_area="Risk Management",
            topic="Risk prioritization and monitoring",
            summary=(
                "Describe how climate-related risks are prioritized, monitored, "
                "and compared with other risks."
            ),
            data_type="Risk-management process",
            evidence_needed=(
                "Risk ranking criteria, monitoring indicators, reporting "
                "frequency, thresholds, and escalation procedures."
            ),
            source_reference="IFRS S2 paragraphs 24–26",
        ),

        make_requirement(
            requirement_id="S2-RM-03",
            content_area="Risk Management",
            topic="Integration with enterprise risk management",
            summary=(
                "Explain how climate-related risk processes are integrated "
                "into the entity's overall risk-management process."
            ),
            data_type="Governance and risk management",
            evidence_needed=(
                "Enterprise risk policy, integrated risk register, ownership, "
                "reporting lines, and risk-governance documentation."
            ),
            source_reference="IFRS S2 paragraph 25",
        ),

        make_requirement(
            requirement_id="S2-MT-01",
            content_area="Metrics and Targets",
            topic="Scope 1 emissions",
            summary=(
                "Disclose gross Scope 1 greenhouse-gas emissions and relevant "
                "measurement information."
            ),
            data_type="Metric",
            evidence_needed=(
                "Annual Scope 1 inventory, unit, reporting boundary, "
                "methodology, factors, and reporting period."
            ),
            source_reference="IFRS S2 paragraph 29",
            quantitative="Yes",
            scope_1="Yes",
        ),

        make_requirement(
            requirement_id="S2-MT-02",
            content_area="Metrics and Targets",
            topic="Scope 2 emissions",
            summary=(
                "Disclose gross Scope 2 greenhouse-gas emissions and relevant "
                "measurement information."
            ),
            data_type="Metric",
            evidence_needed=(
                "Annual Scope 2 inventory, unit, reporting boundary, "
                "methodology, factors, and reporting period."
            ),
            source_reference="IFRS S2 paragraph 29",
            quantitative="Yes",
            scope_2="Yes",
        ),

        make_requirement(
            requirement_id="S2-MT-03",
            content_area="Metrics and Targets",
            topic="Scope 3 emissions",
            summary=(
                "Disclose gross Scope 3 greenhouse-gas emissions and relevant "
                "category, boundary, and measurement information."
            ),
            data_type="Metric",
            evidence_needed=(
                "Scope 3 inventory, categories, exclusions, methodology, "
                "assumptions, factors, boundary, and reporting period."
            ),
            source_reference="IFRS S2 paragraph 29",
            quantitative="Yes",
            scope_3="Yes",
        ),

        make_requirement(
            requirement_id="S2-MT-04",
            content_area="Metrics and Targets",
            topic="Climate-related targets",
            summary=(
                "Disclose climate-related targets, including metric, value, "
                "scope, baseline, target period, and methodology."
            ),
            data_type="Target",
            evidence_needed=(
                "Target value, baseline year, target year, organizational "
                "boundary, emissions scopes, methodology, and validation status."
            ),
            source_reference="IFRS S2 paragraphs 33–36",
            quantitative="Yes",
            time_horizon="Yes",
            scope_1="Yes",
            scope_2="Yes",
            scope_3="Yes",
        ),

        make_requirement(
            requirement_id="S2-MT-05",
            content_area="Metrics and Targets",
            topic="Progress against targets",
            summary=(
                "Disclose performance against climate targets and analyze "
                "changes or trends in performance."
            ),
            data_type="Metric and target",
            evidence_needed=(
                "Current performance, baseline performance, progress percentage, "
                "trend, explanations, and corrective action."
            ),
            source_reference="IFRS S2 paragraphs 33–36",
            quantitative="Yes",
            time_horizon="Yes",
        ),

        make_requirement(
            requirement_id="S2-MT-06",
            content_area="Metrics and Targets",
            topic="Internal carbon price",
            summary=(
                "Disclose whether and how an internal carbon price is used "
                "in decision-making and the price applied."
            ),
            data_type="Metric and methodology",
            evidence_needed=(
                "Internal carbon-price value, currency, application, affected "
                "decisions, assumptions, and governance."
            ),
            source_reference="IFRS S2 paragraph 29",
            quantitative="Yes",
            financial_link="Yes",
        ),

        make_requirement(
            requirement_id="S2-MT-07",
            content_area="Metrics and Targets",
            topic="Climate-linked remuneration",
            summary=(
                "Disclose whether climate-related considerations are incorporated "
                "into executive remuneration and how they affect compensation."
            ),
            data_type="Governance and metric",
            evidence_needed=(
                "Compensation metrics, incentive weighting, performance criteria, "
                "governance approval, and remuneration outcomes."
            ),
            source_reference="IFRS S2 paragraph 29",
            quantitative="Yes",
            financial_link="Yes",
        ),
    ]


def format_sheet(worksheet) -> None:
    """
    Apply readable formatting to a worksheet.
    """

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    for row_number in range(2, worksheet.max_row + 1):
        worksheet.row_dimensions[row_number].height = 85


def create_requirement_sheet(
    workbook: Workbook,
    requirements: list[dict],
) -> None:
    worksheet = workbook.active
    worksheet.title = "Requirement Library"

    worksheet.append(HEADERS)

    for record in requirements:
        worksheet.append(
            [record.get(header) for header in HEADERS]
        )

    widths = {
        "A": 16,
        "B": 12,
        "C": 14,
        "D": 24,
        "E": 32,
        "F": 58,
        "G": 26,
        "H": 16,
        "I": 16,
        "J": 16,
        "K": 18,
        "L": 14,
        "M": 14,
        "N": 14,
        "O": 65,
        "P": 28,
        "Q": 42,
        "R": 28,
        "S": 15,
        "T": 15,
        "U": 18,
        "V": 14,
        "W": 45,
        "X": 16,
        "Y": 45,
    }

    for column_letter, width in widths.items():
        worksheet.column_dimensions[column_letter].width = width

    format_sheet(worksheet)
    add_dropdowns(worksheet)


def add_dropdowns(worksheet) -> None:
    yes_no = DataValidation(
        type="list",
        formula1='"Yes,No,Conditional"',
        allow_blank=False,
    )

    review_status = DataValidation(
        type="list",
        formula1='"Draft,Reviewed,Approved"',
        allow_blank=False,
    )

    worksheet.add_data_validation(yes_no)
    worksheet.add_data_validation(review_status)

    last_row = worksheet.max_row

    for column in ["H", "I", "J", "K", "L", "M", "N", "U", "V"]:
        yes_no.add(f"{column}2:{column}{last_row}")

    review_status.add(f"X2:X{last_row}")


def create_data_dictionary(workbook: Workbook) -> None:
    worksheet = workbook.create_sheet("Data Dictionary")

    rows = [
        ["Field", "Definition", "Example"],
        [
            "Requirement ID",
            "Unique identifier for one disclosure requirement.",
            "S2-GOV-01",
        ],
        [
            "Content Area",
            "IFRS S2 disclosure pillar.",
            "Governance",
        ],
        [
            "Requirement Summary",
            "Plain-language summary of the expected disclosure.",
            "Identify the board committee responsible for climate oversight.",
        ],
        [
            "Evidence Needed",
            "Evidence expected to support readiness or coverage.",
            "Committee charter and responsibility statement.",
        ],
        [
            "Quantitative Required",
            "Whether numerical information is expected.",
            "Yes",
        ],
        [
            "Financial Link Required",
            "Whether the disclosure should connect to financial effects.",
            "Yes",
        ],
        [
            "Review Status",
            "Quality-control stage for the requirement row.",
            "Draft",
        ],
    ]

    for row in rows:
        worksheet.append(row)

    worksheet.column_dimensions["A"].width = 30
    worksheet.column_dimensions["B"].width = 75
    worksheet.column_dimensions["C"].width = 60

    format_sheet(worksheet)


def create_review_instructions(workbook: Workbook) -> None:
    worksheet = workbook.create_sheet("Review Instructions")

    rows = [
        ["Step", "Instruction"],
        [
            1,
            "Verify the Requirement Summary against the official IFRS S2 source.",
        ],
        [
            2,
            "Confirm the Source Reference identifies the correct paragraph.",
        ],
        [
            3,
            "Confirm that Evidence Needed is specific enough for readiness assessment.",
        ],
        [
            4,
            "Check quantitative, financial-link, scenario-analysis, and emissions-scope flags.",
        ],
        [
            5,
            "Change Review Status from Draft to Reviewed only after source verification.",
        ],
        [
            6,
            "Use Approved only after a second quality-control review.",
        ],
        [
            7,
            "Do not place TJX evidence in this workbook. Mapping belongs in a separate table.",
        ],
    ]

    for row in rows:
        worksheet.append(row)

    worksheet.column_dimensions["A"].width = 12
    worksheet.column_dimensions["B"].width = 110

    format_sheet(worksheet)


def create_summary_sheet(
    workbook: Workbook,
    requirements: list[dict],
) -> None:
    worksheet = workbook.create_sheet("Library Summary")

    areas = [
        "Governance",
        "Strategy",
        "Risk Management",
        "Metrics and Targets",
    ]

    rows = [
        ["Metric", "Value"],
        ["Total requirements", len(requirements)],
    ]

    for area in areas:
        count = sum(
            record["Content Area"] == area
            for record in requirements
        )

        rows.append(
            [f"{area} requirements", count]
        )

    rows.extend(
        [
            [
                "Quantitative requirements",
                sum(
                    record["Quantitative Required"] == "Yes"
                    for record in requirements
                ),
            ],
            [
                "Financial-link requirements",
                sum(
                    record["Financial Link Required"] == "Yes"
                    for record in requirements
                ),
            ],
            [
                "Scenario-analysis requirements",
                sum(
                    record["Scenario Analysis Relevant"] == "Yes"
                    for record in requirements
                ),
            ],
            [
                "Scope 3-related requirements",
                sum(
                    record["Scope 3 Relevant"] == "Yes"
                    for record in requirements
                ),
            ],
            ["Draft rows", len(requirements)],
            ["Reviewed rows", 0],
            ["Approved rows", 0],
        ]
    )

    for row in rows:
        worksheet.append(row)

    worksheet.column_dimensions["A"].width = 42
    worksheet.column_dimensions["B"].width = 20

    format_sheet(worksheet)


def main() -> None:
    requirements = build_requirements()

    workbook = Workbook()

    create_requirement_sheet(
        workbook,
        requirements,
    )

    create_data_dictionary(workbook)
    create_review_instructions(workbook)

    create_summary_sheet(
        workbook,
        requirements,
    )

    OUTPUT_FILE.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook.save(OUTPUT_FILE)

    print("=" * 72)
    print("ISSB IFRS S2 REQUIREMENT LIBRARY CREATED")
    print("=" * 72)
    print(f"Requirement rows created: {len(requirements)}")
    print(f"Workbook saved to:\n{OUTPUT_FILE}")
    print("=" * 72)


if __name__ == "__main__":
    main()