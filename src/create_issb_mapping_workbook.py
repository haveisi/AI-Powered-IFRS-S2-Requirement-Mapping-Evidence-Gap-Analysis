import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


PROJECT_ROOT = Path(__file__).resolve().parent.parent

EVIDENCE_FILE = (
    PROJECT_ROOT
    / "05_Evidence_Register"
    / "reviewed_outputs"
    / "TJX_approved_evidence.json"
)

REQUIREMENT_FILE = (
    PROJECT_ROOT
    / "06_Framework_Libraries"
    / "ISSB_S2_Requirement_Library.xlsx"
)

OUTPUT_FILE = (
    PROJECT_ROOT
    / "07_Framework_Mapping"
    / "TJX_ISSB_S2_Mapping_Pilot.xlsx"
)

REQUIREMENT_SHEET = "Requirement Library"


PRIORITY_REQUIREMENT_IDS = {
    "S2-GOV-02",
    "S2-GOV-04",
    "S2-GOV-06",
    "S2-GOV-07",
    "S2-STR-06",
    "S2-STR-07",
    "S2-MT-01",
    "S2-MT-02",
    "S2-MT-04",
    "S2-MT-12",
    "S2-MT-13",
    "S2-MT-14",
}


def load_approved_evidence() -> list[dict[str, Any]]:
    """
    Load human-approved TJX evidence.
    """

    if not EVIDENCE_FILE.exists():
        raise FileNotFoundError(
            f"Approved evidence file not found:\n{EVIDENCE_FILE}"
        )

    with EVIDENCE_FILE.open(
        "r",
        encoding="utf-8",
    ) as file:
        evidence = json.load(file)

    if not isinstance(evidence, list):
        raise ValueError(
            "Approved evidence JSON must contain a list."
        )

    if not evidence:
        raise ValueError(
            "Approved evidence JSON contains no records."
        )

    return evidence


def load_reviewed_requirements() -> list[dict[str, Any]]:
    """
    Load only priority requirements marked Reviewed or Approved.
    """

    if not REQUIREMENT_FILE.exists():
        raise FileNotFoundError(
            f"Requirement workbook not found:\n{REQUIREMENT_FILE}"
        )

    workbook = load_workbook(
        REQUIREMENT_FILE,
        data_only=True,
    )

    if REQUIREMENT_SHEET not in workbook.sheetnames:
        raise ValueError(
            f"Worksheet '{REQUIREMENT_SHEET}' was not found."
        )

    worksheet = workbook[REQUIREMENT_SHEET]

    headers = [
        cell.value
        for cell in worksheet[1]
    ]

    required_headers = {
        "Requirement ID",
        "Content Area",
        "Requirement Topic",
        "Requirement Summary",
        "Evidence Needed",
        "Source Reference",
        "Review Status",
    }

    missing_headers = required_headers - set(headers)

    if missing_headers:
        raise ValueError(
            f"Requirement workbook is missing columns: "
            f"{sorted(missing_headers)}"
        )

    requirements: list[dict[str, Any]] = []

    for row_number in range(
        2,
        worksheet.max_row + 1,
    ):
        values = [
            worksheet.cell(
                row=row_number,
                column=column_number,
            ).value
            for column_number in range(
                1,
                worksheet.max_column + 1,
            )
        ]

        row = dict(
            zip(headers, values)
        )

        requirement_id = row.get(
            "Requirement ID"
        )

        review_status = row.get(
            "Review Status"
        )

        if requirement_id not in PRIORITY_REQUIREMENT_IDS:
            continue

        if review_status not in {
            "Reviewed",
            "Approved",
        }:
            continue

        requirements.append(row)

    if not requirements:
        raise ValueError(
            "No priority requirements are marked Reviewed or Approved. "
            "Review the selected ISSB rows before creating the mapping workbook."
        )

    return requirements


def evidence_value(
    evidence: dict[str, Any],
    field_name: str,
) -> Any:
    """
    Read evidence values while supporting both Excel-style and
    JSON-style field names.
    """

    alternatives = {
        "evidence_id": [
            "Evidence ID",
            "evidence_id",
        ],
        "source_document": [
            "Source Document",
            "source_document",
        ],
        "pdf_page": [
            "PDF Page",
            "pdf_page_number",
        ],
        "topic": [
            "Topic",
            "topic",
        ],
        "claim": [
            "Final Claim",
            "Claim",
            "claim",
        ],
        "exact_quote": [
            "Exact Quote",
            "exact_quote",
        ],
        "evidence_type": [
            "Evidence Type",
            "evidence_type",
        ],
        "metric_name": [
            "Metric Name",
            "metric_name",
        ],
        "metric_value": [
            "Metric Value",
            "metric_value",
        ],
        "metric_unit": [
            "Metric Unit",
            "metric_unit",
        ],
        "reporting_period": [
            "Reporting Period",
            "reporting_period",
        ],
        "geographic_scope": [
            "Geographic Scope",
            "geographic_scope",
        ],
    }

    for key in alternatives[field_name]:
        if key in evidence:
            return evidence.get(key)

    return None


def create_candidate_rows(
    requirements: list[dict[str, Any]],
    evidence_items: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """
    Create blank mapping rows for manual assessment.

    One row represents one possible requirement-evidence relationship.
    """

    rows: list[dict[str, Any]] = []

    for requirement in requirements:
        for evidence in evidence_items:
            rows.append(
                {
                    "Mapping ID": (
                        f"{requirement['Requirement ID']}"
                        f"__{evidence_value(evidence, 'evidence_id')}"
                    ),
                    "Requirement ID": requirement.get(
                        "Requirement ID"
                    ),
                    "Content Area": requirement.get(
                        "Content Area"
                    ),
                    "Requirement Topic": requirement.get(
                        "Requirement Topic"
                    ),
                    "Requirement Summary": requirement.get(
                        "Requirement Summary"
                    ),
                    "Evidence Needed": requirement.get(
                        "Evidence Needed"
                    ),
                    "Source Reference": requirement.get(
                        "Source Reference"
                    ),
                    "Evidence ID": evidence_value(
                        evidence,
                        "evidence_id",
                    ),
                    "Source Document": evidence_value(
                        evidence,
                        "source_document",
                    ),
                    "PDF Page": evidence_value(
                        evidence,
                        "pdf_page",
                    ),
                    "Evidence Topic": evidence_value(
                        evidence,
                        "topic",
                    ),
                    "Final Claim": evidence_value(
                        evidence,
                        "claim",
                    ),
                    "Exact Quote": evidence_value(
                        evidence,
                        "exact_quote",
                    ),
                    "Evidence Type": evidence_value(
                        evidence,
                        "evidence_type",
                    ),
                    "Metric Name": evidence_value(
                        evidence,
                        "metric_name",
                    ),
                    "Metric Value": evidence_value(
                        evidence,
                        "metric_value",
                    ),
                    "Metric Unit": evidence_value(
                        evidence,
                        "metric_unit",
                    ),
                    "Reporting Period": evidence_value(
                        evidence,
                        "reporting_period",
                    ),
                    "Geographic Scope": evidence_value(
                        evidence,
                        "geographic_scope",
                    ),
                    "Mapping Decision": "Not Assessed",
                    "Coverage Level": "Not Assessed",
                    "Evidence Strength": "Not Assessed",
                    "Missing Elements": "",
                    "Reviewer Rationale": "",
                    "Human Review Status": "Pending",
                }
            )

    return rows


def create_workbook(
    requirements: list[dict[str, Any]],
    evidence_items: list[dict[str, Any]],
    mapping_rows: list[dict[str, Any]],
) -> Workbook:
    workbook = Workbook()

    mapping_sheet = workbook.active
    mapping_sheet.title = "Mapping Review"

    headers = list(
        mapping_rows[0].keys()
    )

    mapping_sheet.append(headers)

    for row in mapping_rows:
        mapping_sheet.append(
            [
                row.get(header)
                for header in headers
            ]
        )

    format_mapping_sheet(
        mapping_sheet,
        headers,
    )

    add_mapping_dropdowns(
        mapping_sheet,
        headers,
    )

    create_requirement_sheet(
        workbook,
        requirements,
    )

    create_evidence_sheet(
        workbook,
        evidence_items,
    )

    create_instructions_sheet(
        workbook
    )

    return workbook


def format_mapping_sheet(
    worksheet,
    headers: list[str],
) -> None:
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    widths = {
        "Mapping ID": 34,
        "Requirement ID": 16,
        "Content Area": 22,
        "Requirement Topic": 30,
        "Requirement Summary": 55,
        "Evidence Needed": 55,
        "Source Reference": 24,
        "Evidence ID": 18,
        "Source Document": 42,
        "PDF Page": 12,
        "Evidence Topic": 26,
        "Final Claim": 55,
        "Exact Quote": 75,
        "Evidence Type": 18,
        "Metric Name": 28,
        "Metric Value": 16,
        "Metric Unit": 16,
        "Reporting Period": 18,
        "Geographic Scope": 22,
        "Mapping Decision": 20,
        "Coverage Level": 20,
        "Evidence Strength": 20,
        "Missing Elements": 45,
        "Reviewer Rationale": 55,
        "Human Review Status": 20,
    }

    for column_number, header in enumerate(
        headers,
        start=1,
    ):
        column_letter = worksheet.cell(
            row=1,
            column=column_number,
        ).column_letter

        worksheet.column_dimensions[
            column_letter
        ].width = widths.get(
            header,
            18,
        )

    for row in worksheet.iter_rows(
        min_row=2,
        max_row=worksheet.max_row,
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    for row_number in range(
        2,
        worksheet.max_row + 1,
    ):
        worksheet.row_dimensions[
            row_number
        ].height = 100


def column_letter(
    headers: list[str],
    header_name: str,
) -> str:
    column_number = (
        headers.index(header_name) + 1
    )

    result = ""

    while column_number:
        column_number, remainder = divmod(
            column_number - 1,
            26,
        )

        result = (
            chr(65 + remainder)
            + result
        )

    return result


def add_mapping_dropdowns(
    worksheet,
    headers: list[str],
) -> None:
    mapping_decision = DataValidation(
        type="list",
        formula1=(
            '"Not Assessed,Relevant,'
            'Not Relevant,Duplicate"'
        ),
        allow_blank=False,
    )

    coverage_level = DataValidation(
        type="list",
        formula1=(
            '"Not Assessed,Full,Partial,'
            'Weak,None"'
        ),
        allow_blank=False,
    )

    evidence_strength = DataValidation(
        type="list",
        formula1=(
            '"Not Assessed,Strong,Moderate,Weak"'
        ),
        allow_blank=False,
    )

    review_status = DataValidation(
        type="list",
        formula1=(
            '"Pending,Reviewed,Approved"'
        ),
        allow_blank=False,
    )

    worksheet.add_data_validation(
        mapping_decision
    )

    worksheet.add_data_validation(
        coverage_level
    )

    worksheet.add_data_validation(
        evidence_strength
    )

    worksheet.add_data_validation(
        review_status
    )

    last_row = worksheet.max_row

    mapping_column = column_letter(
        headers,
        "Mapping Decision",
    )

    coverage_column = column_letter(
        headers,
        "Coverage Level",
    )

    strength_column = column_letter(
        headers,
        "Evidence Strength",
    )

    review_column = column_letter(
        headers,
        "Human Review Status",
    )

    mapping_decision.add(
        f"{mapping_column}2:"
        f"{mapping_column}{last_row}"
    )

    coverage_level.add(
        f"{coverage_column}2:"
        f"{coverage_column}{last_row}"
    )

    evidence_strength.add(
        f"{strength_column}2:"
        f"{strength_column}{last_row}"
    )

    review_status.add(
        f"{review_column}2:"
        f"{review_column}{last_row}"
    )


def create_requirement_sheet(
    workbook: Workbook,
    requirements: list[dict[str, Any]],
) -> None:
    worksheet = workbook.create_sheet(
        "Reviewed Requirements"
    )

    headers = [
        "Requirement ID",
        "Content Area",
        "Requirement Topic",
        "Requirement Summary",
        "Evidence Needed",
        "Source Reference",
        "Review Status",
    ]

    worksheet.append(headers)

    for requirement in requirements:
        worksheet.append(
            [
                requirement.get(header)
                for header in headers
            ]
        )

    apply_simple_formatting(
        worksheet
    )


def create_evidence_sheet(
    workbook: Workbook,
    evidence_items: list[dict[str, Any]],
) -> None:
    worksheet = workbook.create_sheet(
        "Approved Evidence"
    )

    headers = [
        "Evidence ID",
        "Source Document",
        "PDF Page",
        "Topic",
        "Final Claim",
        "Exact Quote",
        "Evidence Type",
        "Metric Name",
        "Metric Value",
        "Metric Unit",
        "Reporting Period",
        "Geographic Scope",
    ]

    worksheet.append(headers)

    for evidence in evidence_items:
        worksheet.append(
            [
                evidence_value(
                    evidence,
                    "evidence_id",
                ),
                evidence_value(
                    evidence,
                    "source_document",
                ),
                evidence_value(
                    evidence,
                    "pdf_page",
                ),
                evidence_value(
                    evidence,
                    "topic",
                ),
                evidence_value(
                    evidence,
                    "claim",
                ),
                evidence_value(
                    evidence,
                    "exact_quote",
                ),
                evidence_value(
                    evidence,
                    "evidence_type",
                ),
                evidence_value(
                    evidence,
                    "metric_name",
                ),
                evidence_value(
                    evidence,
                    "metric_value",
                ),
                evidence_value(
                    evidence,
                    "metric_unit",
                ),
                evidence_value(
                    evidence,
                    "reporting_period",
                ),
                evidence_value(
                    evidence,
                    "geographic_scope",
                ),
            ]
        )

    apply_simple_formatting(
        worksheet
    )


def create_instructions_sheet(
    workbook: Workbook,
) -> None:
    worksheet = workbook.create_sheet(
        "Mapping Instructions"
    )

    rows = [
        [
            "Field",
            "How to assess it",
        ],
        [
            "Mapping Decision",
            "Choose Relevant only when the evidence directly supports part or all of the requirement.",
        ],
        [
            "Coverage Level — Full",
            "The evidence addresses all material elements described in the requirement row.",
        ],
        [
            "Coverage Level — Partial",
            "The evidence addresses some important elements but leaves clear missing elements.",
        ],
        [
            "Coverage Level — Weak",
            "The evidence has a loose connection but is insufficient for disclosure readiness.",
        ],
        [
            "Coverage Level — None",
            "The evidence does not support the requirement.",
        ],
        [
            "Evidence Strength",
            "Assess specificity, completeness, quantitative support, reporting period, scope, and source traceability.",
        ],
        [
            "Missing Elements",
            "List exactly what the evidence does not provide, such as oversight frequency, target boundary, methodology, financial effect, or progress.",
        ],
        [
            "Reviewer Rationale",
            "Explain why the evidence is or is not relevant. Do not rely only on shared keywords.",
        ],
        [
            "Human Review Status",
            "Set to Reviewed after completing the mapping assessment. Use Approved only after a second quality check.",
        ],
    ]

    for row in rows:
        worksheet.append(row)

    apply_simple_formatting(
        worksheet
    )


def apply_simple_formatting(
    worksheet,
) -> None:
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:
        max_length = 0

        column_letter_value = (
            column_cells[0].column_letter
        )

        for cell in column_cells:
            value = cell.value

            if value is None:
                continue

            max_length = max(
                max_length,
                min(
                    len(str(value)),
                    70,
                ),
            )

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

        worksheet.column_dimensions[
            column_letter_value
        ].width = max(
            14,
            min(
                max_length + 2,
                70,
            ),
        )


def main() -> None:
    evidence_items = load_approved_evidence()

    requirements = load_reviewed_requirements()

    mapping_rows = create_candidate_rows(
        requirements=requirements,
        evidence_items=evidence_items,
    )

    workbook = create_workbook(
        requirements=requirements,
        evidence_items=evidence_items,
        mapping_rows=mapping_rows,
    )

    OUTPUT_FILE.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook.save(
        OUTPUT_FILE
    )

    print("=" * 72)
    print("TJX ISSB S2 MAPPING PILOT")
    print("=" * 72)

    print(
        f"Reviewed requirements loaded: "
        f"{len(requirements)}"
    )

    print(
        f"Approved evidence items loaded: "
        f"{len(evidence_items)}"
    )

    print(
        f"Candidate mapping rows created: "
        f"{len(mapping_rows)}"
    )

    print(
        f"Workbook saved to:\n{OUTPUT_FILE}"
    )

    print("=" * 72)


if __name__ == "__main__":
    main()